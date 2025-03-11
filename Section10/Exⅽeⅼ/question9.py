from openpyxl import Workbook

# ایجاد یک فایل اکسل جدید
wb = Workbook()
ws = wb.active
ws.title = "اعداد"

# افزودن عنوان ستون
ws.append(["اعداد"])

# لیست اعداد نمونه
numbers = [45, 12, 78, 23, 89, 5, 67, 99, 34, 56]

# افزودن اعداد به اکسل
for num in numbers:
    ws.append([num])

# افزودن فرمول برای پیدا کردن بزرگ‌ترین و کوچک‌ترین عدد
ws["B1"] = "بزرگ‌ترین عدد:"
ws["B2"] = "=MAX(A:A)"  # فرمول اکسل برای پیدا کردن بیشترین مقدار

ws["C1"] = "کوچک‌ترین عدد:"
ws["C2"] = "=MIN(A:A)"  # فرمول اکسل برای پیدا کردن کمترین مقدار

# ذخیره فایل اکسل
file_name = "numbers.xlsx"
wb.save(file_name)

print(f"فایل '{file_name}' با موفقیت ذخیره شد.")
