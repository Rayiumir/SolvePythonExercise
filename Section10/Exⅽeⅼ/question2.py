from openpyxl import Workbook

# ایجاد یک شیء Workbook
wb = Workbook()

# انتخاب برگه فعال
ws = wb.active
ws.title = "اعداد"

# وارد کردن داده‌ها (لیست اعداد)
numbers = [12, 15, 20, 25, 30, 35, 40, 45]

# وارد کردن لیست اعداد به اکسل
for idx, number in enumerate(numbers, start=1):
    ws[f"A{idx}"] = number

# وارد کردن فرمول میانگین در سلول B1
ws["B1"] = "=AVERAGE(A1:A8)"

# ذخیره فایل اکسل
wb.save("numbers_with_average.xlsx")

# پیامی که نشان دهد فایل ذخیره شده است
print("فایل اکسل با موفقیت ذخیره شد.")
