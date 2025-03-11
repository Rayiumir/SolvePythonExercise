from openpyxl import Workbook
from datetime import datetime

# ایجاد یک شیء Workbook
wb = Workbook()

# انتخاب برگه فعال
ws = wb.active
ws.title = "تاریخ تولد"

# وارد کردن نام افراد و تاریخ تولد آنها
people = [
    ("علی", "2005-02-10"),
    ("سارا", "2000-05-22"),
    ("محمد", "2010-11-15"),
    ("مریم", "2003-09-03"),
    ("حسین", "2007-06-19")
]

# وارد کردن داده‌ها به اکسل
ws.append(["نام", "تاریخ تولد"])  # عنوان ستون‌ها
for person in people:
    ws.append(person)

# اضافه کردن فرمول برای محاسبه سن افراد
ws["C1"] = "سن"

# محاسبه سن هر فرد و وارد کردن آن در ستون C
for i in range(2, len(people) + 2):
    birth_date = ws[f"B{i}"].value  # تاریخ تولد فرد
    age = (datetime.now() - datetime.strptime(birth_date, "%Y-%m-%d")).days // 365  # محاسبه سن
    ws[f"C{i}"] = age

# اضافه کردن فرمول برای محاسبه تعداد افراد بزرگتر از ۱۸ سال
ws["D1"] = "افراد بزرگتر از ۱۸ سال"
ws["D2"] = "=COUNTIF(C2:C6, \">18\")"

# ذخیره فایل اکسل
wb.save("people_with_birthdates_and_age.xlsx")

# پیامی که نشان دهد فایل ذخیره شده است
print("فایل اکسل با موفقیت ذخیره شد.")
