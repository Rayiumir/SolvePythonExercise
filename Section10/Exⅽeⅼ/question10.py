from openpyxl import Workbook

# ایجاد یک فایل اکسل جدید
wb = Workbook()
ws = wb.active
ws.title = "فروش‌ها"

# افزودن عنوان ستون‌ها
ws.append(["محصول", "تعداد", "قیمت واحد", "مجموع فروش"])

# داده‌های فروش (نام محصول، تعداد، قیمت واحد)
sales_data = [
    ("لپ‌تاپ", 3, 15000000),
    ("موبایل", 5, 8000000),
    ("هدفون", 10, 1200000),
    ("مانیتور", 2, 6000000),
    ("موس", 8, 500000),
]

# افزودن داده‌ها به اکسل
for row, (product, quantity, price) in enumerate(sales_data, start=2):
    ws[f"A{row}"] = product  # نام محصول
    ws[f"B{row}"] = quantity  # تعداد
    ws[f"C{row}"] = price  # قیمت واحد
    ws[f"D{row}"] = f"=B{row}*C{row}"  # فرمول محاسبه مجموع فروش

# محاسبه مجموع کل فروش
total_row = len(sales_data) + 2  # ردیف بعد از داده‌ها
ws[f"C{total_row}"] = "مجموع کل فروش:"
ws[f"D{total_row}"] = f"=SUM(D2:D{total_row-1})"  # فرمول جمع کل فروش

# ذخیره فایل اکسل
file_name = "sales.xlsx"
wb.save(file_name)

print(f"فایل '{file_name}' با موفقیت ذخیره شد.")
